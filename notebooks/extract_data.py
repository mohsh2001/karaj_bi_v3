import asyncio
import aiohttp
import json
import os
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from playwright.async_api import async_playwright

BASE_URL = "https://jo.opensooq.com/en/cars/cars-for-sale"
OUTPUT_FILE = "opensooq_cars.xlsx"
MAX_CONCURRENT = 10
HEADERS = {
   "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
   "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
   "Accept-Language": "en-US,en;q=0.9",
   "Connection": "keep-alive",
}
COLUMNS = [
   "Title", "Brand", "Model", "Trim", "Year", "Body Type", "Condition", "Config",
   "Mileage", "Price", "Currency", "Seats", "Transmission", "Engine Size",
   "Fuel", "Exterior Color", "Interior Color", "Body Condition", "Paint",
   "Payment Method", "Neighbourhood", "City", "Description", "Link", "Image"
]

async def fetch(session, url):
   try:
       async with session.get(url, headers=HEADERS, timeout=aiohttp.ClientTimeout(total=15)) as resp:
           if resp.status != 200:
               return None
           return await resp.text()
   except Exception:
       return None

def parse_listings(html):
   soup = BeautifulSoup(html, "html.parser")
   listings = []
   for script in soup.find_all("script", type="application/ld+json"):
       try:
           data = json.loads(script.string)
       except (json.JSONDecodeError, TypeError):
           continue
       for node in data.get("@graph", []):
           if node.get("@type") != "ItemList":
               continue
           for entry in node.get("itemListElement", []):
               car = entry.get("item", {})
               listings.append({
                   "Title": car.get("name", ""),
                   "Brand": car.get("brand", {}).get("name", "") if isinstance(car.get("brand"), dict) else car.get("brand", ""),
                   "Model": car.get("model", ""),
                   "Year": car.get("vehicleModelDate", ""),
                   "Body Type": car.get("bodyType", ""),
                   "Config": car.get("vehicleConfiguration", ""),
                   "Mileage": car.get("mileageFromOdometer", {}).get("value", "") if isinstance(car.get("mileageFromOdometer"), dict) else "",
                   "Price": car.get("offers", {}).get("price", ""),
                   "Currency": car.get("offers", {}).get("priceCurrency", ""),
                   "Link": car.get("url", ""),
                   "Image": car.get("image", {}).get("url", "") if isinstance(car.get("image"), dict) else car.get("image", ""),
                   "Description": car.get("description", ""),
               })
   return listings

def parse_detail_html(html):
   details = {}
   if not html:
       return details
   soup = BeautifulSoup(html, "html.parser")
   mapping = {
       "trim": "Trim",
       "number of seats": "Seats",
       "seats": "Seats",
       "transmission": "Transmission",
       "engine size": "Engine Size",
       "fuel": "Fuel",
       "exterior color": "Exterior Color",
       "interior color": "Interior Color",
       "body condition": "Body Condition",
       "paint": "Paint",
       "payment method": "Payment Method",
       "neighborhood": "Neighbourhood",
       "neighbourhood": "Neighbourhood",
       "city": "City",
       "kilometers": "Mileage",
       "mileage": "Mileage",
       "year": "Year",
       "body type": "Body Type",
       "condition": "Condition",
       "car make": "Brand",
       "model": "Model",
       "config": "Config",
       "configuration": "Config",
       "sub category": None,
       "category": None,
       "location on map": None,
   }
   for li in soup.select("section#listingViewBasicInfo li"):
       span = li.find("span")
       a = li.find("a")
       if not span or not a:
           continue
       label = span.get_text(strip=True).lower()
       value = a.get_text(strip=True)
       for key, field in mapping.items():
           if key in label:
               if field and value:
                   details[field] = value
               break
   return details

async def page_worker(queue, context, results):
   page = await context.new_page()
   try:
       while True:
           try:
               idx, car = queue.get_nowait()
           except asyncio.QueueEmpty:
               break
           link = car.get("Link", "")
           if not link:
               continue
           if not link.startswith("http"):
               link = "https://jo.opensooq.com" + link
           try:
               await page.goto(link, wait_until="domcontentloaded", timeout=20000)
               try:
                   view_more = page.locator("section#listingViewBasicInfo button", has_text="View More")
                   await view_more.wait_for(state="visible", timeout=5000)
                   await view_more.click()
                   await page.wait_for_timeout(1000)
               except Exception:
                   pass
               html = await page.content()
               details = parse_detail_html(html)
               car.update(details)
               results[idx] = car
               print(f"  [{idx+1}] OK: {car.get('Title', link)}")
           except Exception as e:
               results[idx] = car
               print(f"  [{idx+1}] FAIL: {link} — {e}")
   finally:
       await page.close()

async def scrape_all():
   all_cars = []
   async with aiohttp.ClientSession() as session:
       page = 1
       while True:
           url = BASE_URL if page == 1 else f"{BASE_URL}?page={page}"
           print(f"Page {page}: {url}")
           html = await fetch(session, url)
           if not html:
               print("  No response, stopping.")
               break
           listings = parse_listings(html)
           if not listings:
               print("  No listings found, stopping.")
               break
           all_cars.extend(listings)
           print(f"  {len(listings)} listings (total: {len(all_cars)})")
           page += 1
   if not all_cars:
       print("No cars found.")
       return
   print(f"\nFetching {len(all_cars)} detail pages with {MAX_CONCURRENT} browser tabs...")
   queue = asyncio.Queue()
   results = {}
   for i, car in enumerate(all_cars):
       queue.put_nowait((i, car))
   async with async_playwright() as pw:
       browser = await pw.chromium.launch(headless=True)
       context = await browser.new_context(
           user_agent=HEADERS["User-Agent"],
           locale="en-US",
       )
       workers = [page_worker(queue, context, results) for _ in range(MAX_CONCURRENT)]
       await asyncio.gather(*workers)
       await context.close()
       await browser.close()
   final_cars = [results[i] for i in sorted(results.keys())]
   save_to_excel(final_cars)

def save_to_excel(cars):
   if os.path.exists(OUTPUT_FILE):
       wb = load_workbook(OUTPUT_FILE)
       ws = wb.active
   else:
       wb = Workbook()
       ws = wb.active
       ws.title = "Cars"
       ws.append(COLUMNS)
   for car in cars:
       ws.append([car.get(col, "") for col in COLUMNS])
   wb.save(OUTPUT_FILE)
   print(f"Saved {len(cars)} rows to {OUTPUT_FILE} (total: {ws.max_row - 1})")

if __name__ == "__main__":
   print(f"Base URL: {BASE_URL}\n")
   asyncio.run(scrape_all())